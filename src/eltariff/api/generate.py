"""API endpoints for generating deployable API code."""

import io
import json
import zipfile
from typing import Any

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pydantic import BaseModel

from ..models.rise_schema import TariffsResponse
from ..services.api_generator import APIGenerator

router = APIRouter(prefix="/api/generate", tags=["generate"])


class GenerateRequest(BaseModel):
    """Request to generate deployment package."""

    tariffs_json: str
    company_name: str
    company_org_no: str


@router.post("/package")
async def generate_package(request: GenerateRequest) -> Response:
    """Generate a complete deployment package as a ZIP file."""
    try:
        # Parse tariffs from JSON
        tariffs_data = json.loads(request.tariffs_json)
        tariffs = TariffsResponse.model_validate(tariffs_data)

        # Generate package
        generator = APIGenerator()
        files = generator.generate_deployment_package(
            tariffs, request.company_name, request.company_org_no
        )

        # Create ZIP file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for filename, content in files.items():
                zip_file.writestr(filename, content)

        zip_buffer.seek(0)

        # Create safe filename
        safe_name = (
            request.company_name.lower()
            .replace(" ", "-")
            .replace("å", "a")
            .replace("ä", "a")
            .replace("ö", "o")
        )

        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}-tariff-api.zip"'
            },
        )
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate: {str(e)}")


@router.post("/preview")
async def preview_package(request: GenerateRequest) -> dict[str, Any]:
    """Preview the generated files without downloading."""
    try:
        # Parse tariffs from JSON
        tariffs_data = json.loads(request.tariffs_json)
        tariffs = TariffsResponse.model_validate(tariffs_data)

        # Generate package
        generator = APIGenerator()
        files = generator.generate_deployment_package(
            tariffs, request.company_name, request.company_org_no
        )

        return {"files": files}
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to preview: {str(e)}")


@router.post("/openapi")
async def generate_openapi(request: GenerateRequest) -> dict[str, Any]:
    """Generate only the OpenAPI specification."""
    try:
        tariffs_data = json.loads(request.tariffs_json)
        tariffs = TariffsResponse.model_validate(tariffs_data)

        generator = APIGenerator()
        return generator.generate_openapi_spec(
            tariffs, request.company_name, request.company_org_no
        )
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate: {str(e)}")


@router.post("/json")
async def generate_json(request: GenerateRequest) -> Response:
    """Download the tariff data as a JSON file (RISE API format)."""
    try:
        tariffs_data = json.loads(request.tariffs_json)
        tariffs = TariffsResponse.model_validate(tariffs_data)

        # Export as formatted JSON
        json_content = tariffs.model_dump_json(by_alias=True, indent=2)

        # Create safe filename
        safe_name = (
            request.company_name.lower()
            .replace(" ", "-")
            .replace("å", "a")
            .replace("ä", "a")
            .replace("ö", "o")
        )

        return Response(
            content=json_content,
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}-tariffer.json"'
            },
        )
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate JSON: {str(e)}")


@router.post("/excel")
async def generate_excel(request: GenerateRequest) -> Response:
    """Generate an Excel file with tariff data for easy sharing."""
    try:
        tariffs_data = json.loads(request.tariffs_json)
        tariffs = TariffsResponse.model_validate(tariffs_data)

        # Create workbook with single flat sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Tariffer"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="017E7A", end_color="017E7A", fill_type="solid")
        section_fill = PatternFill(start_color="E8F5F4", end_color="E8F5F4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Tariff", "Avgiftstyp", "Avgiftsnamn", "Pris ex moms", "Pris ink moms",
            "Enhet", "Period", "Tidsregel", "Beskrivning"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        def format_time_rule(comp) -> str:
            """Format time rules from recurringPeriods."""
            if not comp.recurring_periods:
                return ""
            parts = []
            for rp in comp.recurring_periods:
                for ap in rp.active_periods or []:
                    time_str = ""
                    if ap.from_including and ap.to_excluding:
                        from_t = str(ap.from_including)[:5]
                        to_t = str(ap.to_excluding)[:5]
                        if from_t != "00:00" or to_t != "00:00":
                            time_str = f"{from_t}-{to_t}"

                    day_str = ""
                    if ap.calendar_pattern_references:
                        includes = ap.calendar_pattern_references.include or []
                        if "weekdays" in includes:
                            day_str = "vardagar"
                        elif "weekends" in includes:
                            day_str = "helger"

                    if time_str or day_str:
                        parts.append(f"{time_str} {day_str}".strip())
            return " | ".join(parts) if parts else ""

        def format_period(period: str | None) -> str:
            if period == "P1M":
                return "per månad"
            elif period == "P1Y":
                return "per år"
            elif period == "P1D":
                return "per dag"
            return period or ""

        # Add rows for each tariff
        for tariff in tariffs.tariffs:
            # Fixed prices
            if tariff.fixed_price and tariff.fixed_price.components:
                for comp in tariff.fixed_price.components:
                    ws.append([
                        tariff.name,
                        "Fast avgift",
                        comp.name,
                        float(comp.price.price_ex_vat) if comp.price else 0,
                        float(comp.price.price_inc_vat) if comp.price else 0,
                        "kr",
                        format_period(comp.priced_period),
                        format_time_rule(comp),
                        comp.description or ""
                    ])

            # Energy prices
            if tariff.energy_price and tariff.energy_price.components:
                for comp in tariff.energy_price.components:
                    ws.append([
                        tariff.name,
                        "Energiavgift",
                        comp.name,
                        float(comp.price.price_ex_vat) if comp.price else 0,
                        float(comp.price.price_inc_vat) if comp.price else 0,
                        f"kr/{comp.unit.value}" if comp.unit else "kr/kWh",
                        "",
                        format_time_rule(comp),
                        comp.description or ""
                    ])

            # Power prices
            if tariff.power_price and tariff.power_price.components:
                for comp in tariff.power_price.components:
                    # Add peak info to description
                    desc = comp.description or ""
                    if comp.peak_identification_settings:
                        ps = comp.peak_identification_settings
                        if ps.number_of_peaks_for_average_calculation and ps.number_of_peaks_for_average_calculation > 1:
                            peak_info = f"Snitt av {ps.number_of_peaks_for_average_calculation} toppar"
                            desc = f"{peak_info}. {desc}" if desc else peak_info

                    ws.append([
                        tariff.name,
                        "Effektavgift",
                        comp.name,
                        float(comp.price.price_ex_vat) if comp.price else 0,
                        float(comp.price.price_inc_vat) if comp.price else 0,
                        f"kr/{comp.unit.value}" if comp.unit else "kr/kW",
                        format_period(comp.priced_period) if comp.priced_period else "per månad",
                        format_time_rule(comp),
                        desc
                    ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Create safe filename
        safe_name = (
            request.company_name.lower()
            .replace(" ", "-")
            .replace("å", "a")
            .replace("ä", "a")
            .replace("ö", "o")
        )

        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}-tariffer.xlsx"'
            },
        )
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {str(e)}")
